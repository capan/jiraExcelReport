from openpyxl import load_workbook
import re
import jira
from jira import JIRA
import sys
import os
import tkinter.filedialog


# Jira Server Connection
options = {
    'server': 'YOUR_DOMAIN'}
# Auth
try:
    jira = JIRA(options, basic_auth=('JIRA_USER_NAME', 'JIRA_PASSWORD'))
except BaseException as Be:
    print(Be)
props = jira.application_properties()


regex = re.compile(r'^.*?(?=@)')


user_input = tkinter.filedialog.askopenfilename()
if os.path.exists(user_input):
    workbook = load_workbook(user_input)
else:
    assert os.path.exists(
        user_input), "File couldn't be found at, "+str(user_input)


first_sheet = workbook.get_sheet_names()[0]
worksheet = workbook.get_sheet_by_name(first_sheet)

names = []
emails = []
usernames = []

i = 0
for row in worksheet.iter_rows():
    if row[0].row > 1:
        names.append(row[1].value)
        emails.append(row[2].value)
        usernames.append(row[1].value.lower().replace(' ','.'))
        try:
            jira.add_user(usernames[i], emails[i], directoryId=1, password=None,
                          fullname=names[i], notify=True, active=True, ignore_existing=True)
        except BaseException as Be:
            print ('User Already Exists! : ',usernames[i])
            # print(Be)
        try:
            jira.add_user_to_group(usernames[i], 'USER_GROUP')
        except BaseException as Be:
            print (Be)
        print ('User added: ',usernames)
        i += 1